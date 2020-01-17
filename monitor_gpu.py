#!/usr/bin/env python
# -*- coding: utf-8 -*-

from email import encoders
from email.header import Header
from email.mime.text import MIMEText
from email.utils import parseaddr, formataddr
from email.mime.multipart import MIMEMultipart 
from email.mime.base import MIMEBase
from openpyxl import Workbook

import openpyxl
import os
import numpy as np
import pandas as pd
import time
import subprocess
import logging
import socket
import socks
import requests
import datetime
import smtplib
import openpyxl

logging.basicConfig(
    level=logging.DEBUG,
    format="[%(asctime)s] %(name)s:%(levelname)s: %(message)s"
)

hostname = socket.gethostname()

def _format_addr(s):
    name, addr = parseaddr(s)
    return formataddr((Header(name, 'utf-8').encode(), addr))

'''
@brief 
@param {boolean} is_console - whether export message to console
'''
def monitor_nowtime(is_console):
    used_message_arr = [];

    # 获取进程信息
    t0 = time.time()
    command1 = 'nvidia-smi --query-compute-apps=pid,gpu_bus_id,used_gpu_memory --format=csv,nounits'
    df1 = pd.read_csv(os.popen(command1))
    # 获取GPU对应信息
    command2 = 'nvidia-smi --query-gpu=index,pci.bus_id,memory.total --format=csv,nounits'
    df2 = pd.read_csv(os.popen(command2))
    t1 = time.time()
    # 获取GPU数量
    gpu_num = df2.shape[0]
    
    if is_console :
        logging.info('查询用时:%s' % (t1 - t0))

    # 遍历进程，验证执行人是否正确
    for idx, row in df1.iterrows():
        #print(row)
        pid = row.loc['pid']  # 进程号
        if is_console :
            logging.info('检测进程:%s' % pid)
        # 下列执行命令行方法可以避免 can`t open file 错误
        p1 = subprocess.Popen(["ps", "-eo", "euser,pid"], stdout=subprocess.PIPE)
        p2 = subprocess.Popen(["grep", '\s'+str(row['pid'])], stdin=p1.stdout, stdout=subprocess.PIPE)
        p3 = subprocess.Popen(["cut", "-c", "1-6"], stdin=p2.stdout, stdout=subprocess.PIPE)
        process_user_name = p3.communicate()[0].decode()  # 得到进程执行人名
        names = process_user_name.split('\n')

        p4 = subprocess.Popen(["ps", "aux"], stdout=subprocess.PIPE)
        p5 = subprocess.Popen(["grep", '\s' + str(row['pid']) + '\s'], stdin=p4.stdout, stdout=subprocess.PIPE)
        process_info = p5.communicate()[0]
        
        if not process_info:  # 未查到进程信息：
            continue
            
        for name in names:
            if not name:  # 去掉空字符
                continue
            else:
                process_user_name = name
                break
        if not process_user_name:  # 未匹配到执行人
            continue
        if is_console :
            logging.info('执行人为:%s' % process_user_name)
        user_name = process_user_name[:12]  # ps aux 显示用户名不全。为统一只取前六位
        gpu_index = df2.loc[df2[df2[' pci.bus_id'] == row[' gpu_bus_id']].index.tolist()[0], 'index']
        if is_console :
            logging.info('所用的GPU为:%s' % gpu_index)
        used_memory = row[' used_gpu_memory [MiB]']
        if is_console :
            logging.info('占用显存为：%s MB' % used_memory)
        #print(df2[df2[' pci.bus_id'] == row[' gpu_bus_id']])
        gpu_memory = df2.loc[df2[df2[' pci.bus_id'] == row[' gpu_bus_id']].index.tolist()[0], ' memory.total [MiB]']
        ratio = float(used_memory) / float(gpu_memory) * 100
        if is_console :
            logging.info("占用该gpu显存比率为:%.3f %%" % ratio)

        used_gpu_message = [ process_user_name, ratio ]
        used_message_arr.append(used_gpu_message)
        
        #logging.info('该GPU实际有权使用人：%s' % GPU_RELATION[hostname][gpu_index])
        
        '''
        if user_name not in GPU_RELATION[hostname][gpu_index]:
            # 在错误的GPU上执行了程序
            command3 = 'sudo kill -9 %s' % pid
            os.popen(command3)
            logging.info('杀死进程%s' % pid)
            logging.info('df1:%s' % df1)
            logging.info('df2:%s' % df2)
            logging.info('row:%s' % row)
            send_msg('发现GPU冲突进程%s,进程所属人%s,GPU编号%s,%s' % (pid, user_name, gpu_index,process_info))
        '''

        if is_console :
            logging.info("进程信息：%s" % process_info)

    return used_message_arr

def send_statistics(from_addr,password,to_addr_list,smtp_server,proxy=False):
    if proxy:
        #配置sock5代理
        socks.setdefaultproxy(socks.PROXY_TYPE_SOCKS5, '192.168.*.*', 6666)
    filename = "./log/gpu_usage.txt"
    xlsx_file_name './log/gpu_usage.xlsx'
    wb = Workbook()  #创建xlsx统计表
    ws #当前的sheet
    now_hour = -1;   #当前小时数
    now_minute = -1;  #当前分钟数
    now_day = -1;
    count = 0;
    #key:process_user_name   value:[used_time,used_memory_ratio]
    mess_dict = {}   #在当前这个小时内gpu使用信息字典，包括使用者，以及该人的使用时间和占用内存
    old_time = time.localtime(time.time())
    old_timestamp = time.asctime(old_time)
    while True :
        local_time = time.localtime(time.time())

        if now_hour < 0 :
            now_hour = local_time.tm_hour
            # now_hour = local_time.tm_min
        if now_minute < 0 :
            now_minute = local_time.tm_min
            # now_minute = local_time.tm_sec
        if now_day < 0 :
            now_day = local_time.tm_wday

        #隔一分钟获取一次gpu信息
        if now_minute == local_time.tm_min :
        # if now_minute == local_time.tm_sec:
            continue
        else :
            now_minute = local_time.tm_min
            # now_minute = local_time.tm_sec

        if now_day != local_time.tm_wday:
            #按天创建新的excel sheet
            ws = wb.create_sheet('day' + now_day, 0)
            ws['A1']='姓名'
            ws['B1']='时刻'
            ws['C1']='使用时间'
            ws['D1']='gpu占比'                        

        if now_hour != local_time.tm_hour :
        # if now_hour != local_time.tm_min:
            #把一个小时内gpu使用情况写入文件,txt和excel
            with open(filename, 'a') as f:
                #timestamp = time.asctime(local_time)
                now_timestamp = time.asctime(local_time)
                f.write(old_timestamp + "  ~  " + now_timestamp + "\n")
                print("count:" + str(count))
                print("mess_dict" + str(mess_dict))
                for key,value in mess_dict.items() :
                    f.write(key+ ": 使用时间" + str(value[0]) + "mins  占用gpu百分比：" + str(value[1]/(count * gpu_num)) + "%\n")
                f.write("\n")
                f.close()
                old_timestamp = now_timestamp
            ws.cell(row=now_hour+1,column=1,value=key)
            ws.cell(row=now_hour+1,column=2,value=now_hour+1)
            ws.cell(row=now_hour+1,column=3,value=value[0])
            ws.cell(row=now_hour+1,column=4,value=value[1]/(count * gpu_num))

            now_hour = local_time.tm_hour
            # now_hour = local_time.tm_min
            mess_dict = {}
            count = 0                

        # 每周发送一次实验室GPU使用情况统计信息
        # if datetime.datetime.now().weekday() == 3:
        #     to_addr = ','.join(['zhangqingrui@bupt.edu.cn','wdwang@bupt.edu.cn'])
        #     # 邮件对象:
        #     msg = MIMEMultipart()
        #     host_name = os.popen("hostname").read().strip('\n')
        #     msg['From'] = _format_addr(host_name + ' <%s>' % from_addr)
        #     msg['To'] = to_addr
        #     msg['Subject'] = Header('服务器显卡使用情况', 'utf-8').encode()

        #     # 邮件正文是MIMEText:
        #     msg.attach(MIMEText(host_name, 'plain', 'utf-8'))

        #     # 添加附件就是加上一个MIMEBase，从本地读取文件:
        #     with open('./log/gpu_usage.txt', 'rb') as f:
        #         # 设置附件的MIME和文件名
        #         mime = MIMEBase('text', 'plain', filename='gpu_usage.txt')
        #         # 加上必要的头信息:
        #         mime.add_header('Content-Disposition', 'attachment', filename='gpu_usage.txt')
        #         mime.add_header('Content-ID', '<0>')
        #         mime.add_header('X-Attachment-Id', '0')
        #         # 把附件的内容读进来:
        #         mime.set_payload(f.read())
        #         # 用Base64编码:
        #         encoders.encode_base64(mime)
        #         # 添加到MIMEMultipart:
        #         msg.attach(mime)
        #     with open('./log/gpu_usage.xlsx', 'rb') as f:
        #         # 设置附件的MIME和文件名
        #         mime = MIMEBase('text', 'plain', filename='gpu_usage.xlsx')
        #         # 加上必要的头信息:
        #         mime.add_header('Content-Disposition', 'attachment', filename='gpu_usage.xlsx')
        #         mime.add_header('Content-ID', '<0>')
        #         mime.add_header('X-Attachment-Id', '0')
        #         # 把附件的内容读进来:
        #         mime.set_payload(f.read())
        #         # 用Base64编码:
        #         encoders.encode_base64(mime)
        #         # 添加到MIMEMultipart:
        #         msg.attach(mime)        
        #     server = smtplib.SMTP(smtp_server, 25)
        #     server.set_debuglevel(1)
        #     server.login(from_addr, password)
        #     server.sendmail(from_addr, msg['to'].split(','), msg.as_string())
        #     server.quit()
        #     os.system("rm ./log/gpu_usage.txt")
        #     os.system("rm ./log/gpu_usage.xlsx")
                

        count += 1
        used_message_arr = monitor_nowtime(False)
        #used_message:[process_user_name, used_memory_ratio]
        for used_message in used_message_arr :
            if used_message[0] not in mess_dict.keys() :
                mess_dict[used_message[0]] = [1, used_message[1]]
            else :
                tmp_arr = mess_dict[used_message[0]]
                if tmp_arr[0] < count : #  避免用户使用时间重复增加
                    tmp_arr[0] += 1; # 使用时间+1
                tmp_arr[1] += used_message[1] #占用内存比率
                mess_dict[used_message[0]] = tmp_arr

if __name__=="__main__":
    from_addr = 'bupt_zqr@163.com'
    password = 'zqr035036'
    to_addr_list = ['zhangqingrui@bupt.edu.cn','wdwang@bupt.edu.cn']
    smtp_server = 'smtp.163.com'
    is_proxy = False
    send_statistics(from_addr,password,to_addr_list,smtp_server,is_proxy);
